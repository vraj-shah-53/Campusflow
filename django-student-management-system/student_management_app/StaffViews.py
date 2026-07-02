from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib import messages
from django.core.files.storage import FileSystemStorage #To upload Profile Picture
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.core import serializers
import json


from student_management_app.models import CustomUser, Staffs, Courses, Subjects, Students, SessionYearModel, Attendance, AttendanceReport, LeaveReportStaff, FeedBackStaffs, StudentResult


def staff_home(request):
    # Fetching All Students under Staff

    subjects = Subjects.objects.filter(staff_id=request.user.id)
    course_id_list = []
    for subject in subjects:
        course = Courses.objects.get(id=subject.course_id.id)
        course_id_list.append(course.id)
    
    final_course = []
    # Removing Duplicate Course Id
    for course_id in course_id_list:
        if course_id not in final_course:
            final_course.append(course_id)
    
    students_count = Students.objects.filter(course_id__in=final_course).count()
    subject_count = subjects.count()

    # Fetch All Attendance Count
    attendance_count = Attendance.objects.filter(subject_id__in=subjects).count()
    # Fetch All Approve Leave
    staff = Staffs.objects.get(admin=request.user.id)
    leave_count = LeaveReportStaff.objects.filter(staff_id=staff.id, leave_status=1).count()

    #Fetch Attendance Data by Subjects
    subject_list = []
    attendance_list = []
    for subject in subjects:
        attendance_count1 = Attendance.objects.filter(subject_id=subject.id).count()
        subject_list.append(subject.subject_name)
        attendance_list.append(attendance_count1)

    # Fetch student attendance data subject-wise
    subject_charts_data = []
    for subject in subjects:
        course_students = Students.objects.filter(course_id=subject.course_id).order_by('admin__username')
        student_names = []
        present_counts = []
        absent_counts = []
        
        for student in course_students:
            p_count = AttendanceReport.objects.filter(
                status=True, 
                student_id=student.id, 
                attendance_id__subject_id=subject.id
            ).count()
            a_count = AttendanceReport.objects.filter(
                status=False, 
                student_id=student.id, 
                attendance_id__subject_id=subject.id
            ).count()
            
            student_names.append(student.admin.first_name + " " + student.admin.last_name)
            present_counts.append(p_count)
            absent_counts.append(a_count)
            
        subject_charts_data.append({
            "subject_id": subject.id,
            "subject_name": subject.subject_name,
            "student_list": student_names,
            "attendance_present_list": present_counts,
            "attendance_absent_list": absent_counts
        })

    context={
        "students_count": students_count,
        "attendance_count": attendance_count,
        "leave_count": leave_count,
        "subject_count": subject_count,
        "subject_list": subject_list,
        "attendance_list": attendance_list,
        "subject_charts_data": subject_charts_data
    }
    return render(request, "staff_template/staff_home_template.html", context)



def staff_take_attendance(request):
    subjects = Subjects.objects.filter(staff_id=request.user.id)
    session_years = SessionYearModel.objects.filter(admin_creator=request.user.staffs.admin_creator)
    context = {
        "subjects": subjects,
        "session_years": session_years
    }
    return render(request, "staff_template/take_attendance_template.html", context)


def staff_apply_leave(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    leave_data = LeaveReportStaff.objects.filter(staff_id=staff_obj)
    context = {
        "leave_data": leave_data
    }
    return render(request, "staff_template/staff_apply_leave_template.html", context)


def staff_apply_leave_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('staff_apply_leave')
    else:
        leave_date = request.POST.get('leave_date')
        leave_message = request.POST.get('leave_message')

        if not leave_date:
            messages.error(request, "Please select a leave date.")
            return redirect('staff_apply_leave')

        staff_obj = Staffs.objects.get(admin=request.user.id)
        try:
            leave_report = LeaveReportStaff(staff_id=staff_obj, leave_date=leave_date, leave_message=leave_message, leave_status=0)
            leave_report.save()
            messages.success(request, "Applied for Leave.")
            return redirect('staff_apply_leave')
        except:
            messages.error(request, "Failed to Apply Leave")
            return redirect('staff_apply_leave')


def staff_feedback(request):
    staff_obj = Staffs.objects.get(admin=request.user.id)
    feedback_data = FeedBackStaffs.objects.filter(staff_id=staff_obj)
    context = {
        "feedback_data":feedback_data
    }
    return render(request, "staff_template/staff_feedback_template.html", context)


def staff_feedback_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method.")
        return redirect('staff_feedback')
    else:
        feedback = request.POST.get('feedback_message')
        staff_obj = Staffs.objects.get(admin=request.user.id)

        try:
            add_feedback = FeedBackStaffs(staff_id=staff_obj, feedback=feedback, feedback_reply="")
            add_feedback.save()
            messages.success(request, "Feedback Sent.")
            return redirect('staff_feedback')
        except:
            messages.error(request, "Failed to Send Feedback.")
            return redirect('staff_feedback')


# WE don't need csrf_token when using Ajax
@csrf_exempt
def get_students(request):
    # Getting Values from Ajax POST 'Fetch Student'
    subject_id = request.POST.get("subject")
    session_year = request.POST.get("session_year")

    # Students enroll to Course, Course has Subjects
    # Getting all data from subject model based on subject_id
    subject_model = Subjects.objects.get(id=subject_id)

    session_model = SessionYearModel.objects.get(id=session_year)

    students = Students.objects.filter(course_id=subject_model.course_id, session_year_id=session_model)

    # Only Passing Student Id and Student Name Only
    list_data = []

    for student in students:
        data_small={"id":student.admin.id, "name":student.admin.first_name+" "+student.admin.last_name, "username":student.admin.username}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)




@csrf_exempt
def save_attendance_data(request):
    # Get Values from Staf Take Attendance form via AJAX (JavaScript)
    # Use getlist to access HTML Array/List Input Data
    student_ids = request.POST.get("student_ids")
    subject_id = request.POST.get("subject_id")
    attendance_date = request.POST.get("attendance_date")
    session_year_id = request.POST.get("session_year_id")

    subject_model = Subjects.objects.get(id=subject_id)
    session_year_model = SessionYearModel.objects.get(id=session_year_id)

    json_student = json.loads(student_ids)
    # print(dict_student[0]['id'])

    # print(student_ids)
    try:
        # First Attendance Data is Saved on Attendance Model
        attendance = Attendance(subject_id=subject_model, attendance_date=attendance_date, session_year_id=session_year_model)
        attendance.save()

        for stud in json_student:
            # Attendance of Individual Student saved on AttendanceReport Model
            student = Students.objects.get(admin=stud['id'])
            attendance_report = AttendanceReport(student_id=student, attendance_id=attendance, status=stud['status'])
            attendance_report.save()
        return HttpResponse("OK")
    except:
        return HttpResponse("Error")




def staff_update_attendance(request):
    subjects = Subjects.objects.filter(staff_id=request.user.id)
    session_years = SessionYearModel.objects.filter(admin_creator=request.user.staffs.admin_creator)
    context = {
        "subjects": subjects,
        "session_years": session_years
    }
    return render(request, "staff_template/update_attendance_template.html", context)

@csrf_exempt
def get_attendance_dates(request):
    

    # Getting Values from Ajax POST 'Fetch Student'
    subject_id = request.POST.get("subject")
    session_year = request.POST.get("session_year_id")

    # Students enroll to Course, Course has Subjects
    # Getting all data from subject model based on subject_id
    subject_model = Subjects.objects.get(id=subject_id)

    session_model = SessionYearModel.objects.get(id=session_year)

    # students = Students.objects.filter(course_id=subject_model.course_id, session_year_id=session_model)
    attendance = Attendance.objects.filter(subject_id=subject_model, session_year_id=session_model)

    # Only Passing Student Id and Student Name Only
    list_data = []

    for attendance_single in attendance:
        data_small={"id":attendance_single.id, "attendance_date":str(attendance_single.attendance_date), "session_year_id":attendance_single.session_year_id.id}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)


@csrf_exempt
def get_attendance_student(request):
    # Getting Values from Ajax POST 'Fetch Student'
    attendance_date = request.POST.get('attendance_date')
    attendance = Attendance.objects.get(id=attendance_date)

    attendance_data = AttendanceReport.objects.filter(attendance_id=attendance)
    # Only Passing Student Id and Student Name Only
    list_data = []

    for student in attendance_data:
        data_small={"id":student.student_id.admin.id, "name":student.student_id.admin.first_name+" "+student.student_id.admin.last_name, "status":student.status}
        list_data.append(data_small)

    return JsonResponse(json.dumps(list_data), content_type="application/json", safe=False)


@csrf_exempt
def update_attendance_data(request):
    student_ids = request.POST.get("student_ids")

    attendance_date = request.POST.get("attendance_date")
    attendance = Attendance.objects.get(id=attendance_date)

    json_student = json.loads(student_ids)

    try:
        
        for stud in json_student:
            # Attendance of Individual Student saved on AttendanceReport Model
            student = Students.objects.get(admin=stud['id'])

            attendance_report = AttendanceReport.objects.get(student_id=student, attendance_id=attendance)
            attendance_report.status=stud['status']

            attendance_report.save()
        return HttpResponse("OK")
    except:
        return HttpResponse("Error")


def staff_profile(request):
    user = CustomUser.objects.get(id=request.user.id)
    staff = Staffs.objects.get(admin=user)

    context={
        "user": user,
        "staff": staff
    }
    return render(request, 'staff_template/staff_profile.html', context)


def staff_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('staff_profile')
    else:
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        address = request.POST.get('address')

        profile_pic_url = None
        if len(request.FILES) != 0:
            profile_pic = request.FILES['profile_pic']
            fs = FileSystemStorage()
            filename = fs.save(profile_pic.name, profile_pic)
            profile_pic_url = filename

        try:
            customuser = CustomUser.objects.get(id=request.user.id)
            customuser.first_name = first_name
            customuser.last_name = last_name
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()

            staff = Staffs.objects.get(admin=customuser.id)
            staff.address = address
            if profile_pic_url != None:
                staff.profile_pic = profile_pic_url
            staff.save()

            messages.success(request, "Profile Updated Successfully")
            return redirect('staff_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('staff_profile')



def staff_add_result(request):
    subjects = Subjects.objects.filter(staff_id=request.user.id)
    session_years = SessionYearModel.objects.filter(admin_creator=request.user.staffs.admin_creator)
    context = {
        "subjects": subjects,
        "session_years": session_years,
    }
    return render(request, "staff_template/add_result_template.html", context)


def staff_add_result_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('staff_add_result')
    else:
        student_admin_id = request.POST.get('student_list')
        assignment_marks = request.POST.get('assignment_marks')
        exam_marks = request.POST.get('exam_marks')
        total_marks = request.POST.get('total_marks', 100)
        subject_id = request.POST.get('subject')

        student_obj = Students.objects.get(admin=student_admin_id)
        subject_obj = Subjects.objects.get(id=subject_id)

        try:
            # Check if Students Result Already Exists or not
            check_exist = StudentResult.objects.filter(subject_id=subject_obj, student_id=student_obj).exists()
            if check_exist:
                result = StudentResult.objects.get(subject_id=subject_obj, student_id=student_obj)
                result.subject_assignment_marks = assignment_marks
                result.subject_exam_marks = exam_marks
                result.subject_total_marks = total_marks
                result.save()
                messages.success(request, "Result Updated Successfully!")
                return redirect('staff_add_result')
            else:
                result = StudentResult(student_id=student_obj, subject_id=subject_obj, subject_exam_marks=exam_marks, subject_assignment_marks=assignment_marks, subject_total_marks=total_marks)
                result.save()
                messages.success(request, "Result Added Successfully!")
                return redirect('staff_add_result')
        except:
            messages.error(request, "Failed to Add Result!")
            return redirect('staff_add_result')


def staff_export_attendance(request):
    if request.method != "POST":
        subjects = Subjects.objects.filter(staff_id=request.user.id)
        session_years = SessionYearModel.objects.filter(admin_creator=request.user.staffs.admin_creator)
        context = {
            "subjects": subjects,
            "session_years": session_years,
        }
        return render(request, "staff_template/export_attendance_template.html", context)
    else:
        subject_id = request.POST.get('subject')
        session_year_id = request.POST.get('session_year')

        try:
            subject = Subjects.objects.get(id=subject_id)
            session_year = SessionYearModel.objects.get(id=session_year_id)
        except Exception:
            messages.error(request, "Invalid Subject or Session Year selected!")
            return redirect('staff_export_attendance')

        # Fetch students for this course and session
        students = Students.objects.filter(course_id=subject.course_id, session_year_id=session_year).order_by('admin__username')
        # Fetch all attendance records for this subject and session
        attendance_list = Attendance.objects.filter(subject_id=subject, session_year_id=session_year).order_by('attendance_date')

        if not students.exists():
            messages.error(request, "No students found for this subject and session!")
            return redirect('staff_export_attendance')

        # Create workbook
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"

        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True

        # Styles
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_normal = Font(name="Calibri", size=11)
        font_bold = Font(name="Calibri", size=11, bold=True)
        
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Deep navy
        fill_present = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft green
        fill_absent = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft red
        fill_summary = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Light gray
        
        align_center = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        double_bottom_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='double', color='000000')
        )

        # Header Row
        headers = ["Date"]
        for student in students:
            headers.append(f"{student.admin.username}")
        
        ws.append(headers)
        
        # Style Header
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
            
        # Keep track of counts for present/absent
        student_counts = {student.id: [0, 0] for student in students}

        # Populate rows
        row_num = 2
        for att in attendance_list:
            date_str = att.attendance_date.strftime('%d-%m-%Y')
            row_data = [date_str]
            
            # Fetch all attendance reports for this attendance instance
            att_reports = {report.student_id_id: report.status for report in AttendanceReport.objects.filter(attendance_id=att)}
            
            for student in students:
                is_present = att_reports.get(student.id, False)
                if is_present:
                    status_str = "present"
                    student_counts[student.id][0] += 1
                else:
                    status_str = "absent"
                    student_counts[student.id][1] += 1
                row_data.append(status_str)
                
            ws.append(row_data)
            
            # Style data row
            cell_date = ws.cell(row=row_num, column=1)
            cell_date.font = font_normal
            cell_date.alignment = align_center
            cell_date.border = thin_border
            
            for col_idx in range(2, len(row_data) + 1):
                cell_status = ws.cell(row=row_num, column=col_idx)
                status_val = row_data[col_idx - 1]
                cell_status.font = font_normal
                cell_status.alignment = align_center
                cell_status.border = thin_border
                if status_val == "present":
                    cell_status.fill = fill_present
                else:
                    cell_status.fill = fill_absent
            
            row_num += 1

        # Summary Row: Total Present
        present_row = ["total present :"]
        for student in students:
            present_row.append(student_counts[student.id][0])
        ws.append(present_row)
        
        # Style Total Present
        for col_idx in range(1, len(present_row) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_bold
            cell.fill = fill_summary
            cell.alignment = align_center
            cell.border = thin_border
        row_num += 1

        # Summary Row: Total Absent
        absent_row = ["total absent :"]
        for student in students:
            absent_row.append(student_counts[student.id][1])
        ws.append(absent_row)
        
        # Style Total Absent
        for col_idx in range(1, len(absent_row) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_bold
            cell.fill = fill_summary
            cell.alignment = align_center
            cell.border = double_bottom_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Build dynamic filename
        subj_name_clean = "".join(c for c in subject.subject_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
        session_name_clean = f"{session_year.session_start_year.strftime('%Y')}-{session_year.session_end_year.strftime('%Y')}"
        filename = f"attendance_report_{subj_name_clean}_{session_name_clean}.xlsx"

        # Prepare HTTP response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


def staff_upload_assignment(request):
    from student_management_app.models import StudentAssignment, NotificationStudent

    if request.method != "POST":
        subjects = Subjects.objects.filter(staff_id=request.user.id)
        session_years = SessionYearModel.objects.filter(admin_creator=request.user.staffs.admin_creator)
        context = {
            "subjects": subjects,
            "session_years": session_years,
        }
        return render(request, "staff_template/upload_assignment_template.html", context)
    else:
        subject_id = request.POST.get('subject')
        session_year_id = request.POST.get('session_year')
        assignment_title = request.POST.get('assignment_title')
        assignment_description = request.POST.get('assignment_description')
        assignment_file = request.FILES.get('assignment_file')

        if not assignment_file:
            messages.error(request, "Please upload a valid file.")
            return redirect('staff_upload_assignment')

        try:
            subject_obj = Subjects.objects.get(id=subject_id)
            session_year_obj = SessionYearModel.objects.get(id=session_year_id)
            
            # Save StudentAssignment
            assignment = StudentAssignment(
                subject_id=subject_obj,
                session_year_id=session_year_obj,
                assignment_title=assignment_title,
                assignment_description=assignment_description,
                assignment_file=assignment_file
            )
            assignment.save()

            # Fetch all students enrolled under this course and session year
            students = Students.objects.filter(course_id=subject_obj.course_id, session_year_id=session_year_obj)
            
            # Trigger pop-up notifications for all of them
            for student in students:
                NotificationStudent.objects.create(
                    student_id=student,
                    message=f"New assignment '{assignment_title}' uploaded for {subject_obj.subject_name}!"
                )

            messages.success(request, "Assignment uploaded successfully and students notified!")
            return redirect('staff_upload_assignment')
        except Exception as e:
            messages.error(request, f"Failed to upload assignment: {str(e)}")
            return redirect('staff_upload_assignment')
